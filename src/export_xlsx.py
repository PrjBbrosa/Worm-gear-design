"""Export cycle curve data to Excel."""

import numpy as np

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None


def export_cycle_xlsx(path, inputs, steel, wheel, res):
    if Workbook is None:
        raise ImportError("openpyxl is required for Excel export.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Cycle Curves"

    headers = ["phi_rad", "phi_deg", "p_contact_MPa", "sigma_root_MPa",
               "T2_Nm", "eta", "Nc_proxy"]
    ws.append(headers)

    phi = res["phi"]
    deg = phi * 180.0 / np.pi
    for i in range(len(phi)):
        ws.append([
            float(phi[i]),
            float(deg[i]),
            float(res["p_contact_MPa"][i]),
            float(res["sigma_root_MPa"][i]),
            float(res["T2_Nm"][i]),
            float(res["eta"][i]),
            float(res["Nc_proxy"][i]),
        ])

    # Meta sheet
    ws2 = wb.create_sheet("Meta")
    ws2.append(["Key", "Value"])
    for k, v in res["meta"].items():
        ws2.append([str(k), str(v)])

    # Inputs sheet
    ws3 = wb.create_sheet("Inputs")
    ws3.append(["Key", "Value"])
    for k, v in inputs.items():
        ws3.append([str(k), str(v)])

    wb.save(path)
